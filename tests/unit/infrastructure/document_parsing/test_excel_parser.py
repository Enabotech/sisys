"""Excel 文档解析器单元测试

TDD 红阶段：测试 ExcelParser 的多 Sheet 表格提取、空 Sheet 跳过、公式计算值、空文件拒绝、旧版 XLS 拒绝。
使用 openpyxl 创建 fixture XLSX 文件，避免依赖外部文件。
"""

from __future__ import annotations

import os
import tempfile


def _create_xlsx_with_sheets(sheets: dict[str, list[list[str]]]) -> str:
    """创建多 Sheet XLSX fixture

    Args:
        sheets: {sheet_name: [[cell, ...], ...]}
    """
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            assert ws is not None
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def _create_xlsx_with_empty_sheet() -> str:
    """创建含空 Sheet 的 XLSX fixture"""
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "数据表"
    ws1.append(["列A", "列B"])
    ws1.append(["值1", "值2"])
    wb.create_sheet(title="空表")  # 空 Sheet，无数据

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def _create_empty_xlsx() -> str:
    """创建无数据的空 XLSX fixture（仅默认空 Sheet）"""
    from openpyxl import Workbook

    wb = Workbook()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIME_XLS = "application/vnd.ms-excel"


class TestExcelParserCreation:
    """ExcelParser 构造和基本功能测试"""

    def test_create_parser(self) -> None:
        """验证 ExcelParser 可以正常实例化"""
        from src.infrastructure.document_parsing.excel_parser import ExcelParser

        parser = ExcelParser()
        assert parser is not None

    def test_parser_implements_document_parser_port(self) -> None:
        """验证 ExcelParser 满足 DocumentParserPort 协议"""
        from src.domain.ports.document_parser import DocumentParserPort
        from src.infrastructure.document_parsing.excel_parser import ExcelParser

        parser = ExcelParser()
        assert isinstance(parser, DocumentParserPort)


class TestExcelParserMultiSheet:
    """多 Sheet 解析测试"""

    def test_parse_multi_sheet_xlsx(self) -> None:
        """多 Sheet XLSX 每个 Sheet 独立输出为 ParsedTable"""
        from src.infrastructure.document_parsing.excel_parser import ExcelParser

        sheets = {
            "市场分析": [["维度", "目标"], ["市场", "增长20%"]],
            "技术评估": [["项目", "状态"], ["迁移上云", "进行中"]],
        }
        path = _create_xlsx_with_sheets(sheets)
        try:
            parser = ExcelParser()
            result = parser.parse(path, MIME_XLSX)

            assert result.is_completed(), f"解析应成功，实际: {result.error_message}"
            all_tables = [t for p in result.pages for t in p.tables]
            assert len(all_tables) == 2, f"应 2 个 Sheet 表格，实际: {len(all_tables)}"

            # 验证 sheet_name 在 metadata 中
            table = all_tables[0]
            assert "sheet_name" in table.metadata, f"metadata 应包含 sheet_name，实际: {table.metadata}"
            assert table.metadata["sheet_name"] == "市场分析"

            table2 = all_tables[1]
            assert table2.metadata["sheet_name"] == "技术评估"
        finally:
            os.unlink(path)

    def test_empty_sheet_skipped(self) -> None:
        """空 Sheet 应跳过，不生成空表格"""
        from src.infrastructure.document_parsing.excel_parser import ExcelParser

        path = _create_xlsx_with_empty_sheet()
        try:
            parser = ExcelParser()
            result = parser.parse(path, MIME_XLSX)

            assert result.is_completed()
            all_tables = [t for p in result.pages for t in p.tables]
            # 空表 Sheet 应被跳过，仅 1 个表格
            assert len(all_tables) == 1, f"空 Sheet 应跳过，实际表格数: {len(all_tables)}"
            assert all_tables[0].metadata["sheet_name"] == "数据表"
        finally:
            os.unlink(path)


class TestExcelParserDataExtraction:
    """数据提取测试"""

    def test_parse_single_sheet_xlsx(self) -> None:
        """单 Sheet XLSX 正确提取数据"""
        from src.infrastructure.document_parsing.excel_parser import ExcelParser

        sheets = {"Sheet1": [["姓名", "部门"], ["张三", "技术部"]]}
        path = _create_xlsx_with_sheets(sheets)
        try:
            parser = ExcelParser()
            result = parser.parse(path, MIME_XLSX)

            assert result.is_completed()
            all_tables = [t for p in result.pages for t in p.tables]
            assert len(all_tables) == 1

            table = all_tables[0]
            assert table.rows[0] == ["姓名", "部门"]
            assert table.rows[1] == ["张三", "技术部"]
        finally:
            os.unlink(path)


class TestExcelParserEmptyDocument:
    """空文档检测测试"""

    def test_empty_xlsx_returns_failed(self) -> None:
        """空 XLSX 返回 failed"""
        from src.infrastructure.document_parsing.excel_parser import ExcelParser

        path = _create_empty_xlsx()
        try:
            parser = ExcelParser()
            result = parser.parse(path, MIME_XLSX)

            assert result.parse_status == "failed", f"空 XLSX 应返回 failed，实际: {result.parse_status}"
            assert result.error_message is not None
        finally:
            os.unlink(path)


class TestExcelParserLegacyFormatRejection:
    """旧版 XLS 格式拒绝测试"""

    def test_xls_mime_returns_failed(self) -> None:
        """旧版 XLS MIME 返回 failed 并建议转换"""
        from src.infrastructure.document_parsing.excel_parser import ExcelParser

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xls")
        tmp.write(b"not a valid xlsx")
        tmp.close()
        try:
            parser = ExcelParser()
            result = parser.parse(tmp.name, MIME_XLS)

            assert result.parse_status == "failed"
            assert result.error_message is not None
            assert "XLSX" in result.error_message, f"错误信息应建议转换为 XLSX，实际: {result.error_message}"
        finally:
            os.unlink(tmp.name)

    def test_corrupt_xlsx_returns_failed(self) -> None:
        """损坏 XLSX 返回 failed"""
        from src.infrastructure.document_parsing.excel_parser import ExcelParser

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(b"this is not a valid xlsx zip file")
        tmp.close()
        try:
            parser = ExcelParser()
            result = parser.parse(tmp.name, MIME_XLSX)

            assert result.parse_status == "failed"
            assert result.error_message is not None
        finally:
            os.unlink(tmp.name)
