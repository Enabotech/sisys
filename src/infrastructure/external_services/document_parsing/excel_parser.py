"""Excel 文档解析器

使用 openpyxl 提取 XLSX 多 Sheet 表格内容的解析器实现。
"""

from __future__ import annotations

import logging
import os
import uuid
from datetime import UTC, datetime

from src.domain.ports.document_parser import DocumentParserPort
from src.domain.value_objects.parsed_document import (
    ParsedDocument,
    ParsedPage,
    ParsedTable,
)
from src.infrastructure.external_services.document_parsing._limits import MAX_XLSX_BYTES

logger = logging.getLogger(__name__)


class ExcelParser(DocumentParserPort):
    """Excel 文档解析器

    使用 openpyxl.load_workbook(read_only=True, data_only=True) 提取表格内容，支持：
    - 多 Sheet 独立输出为 ParsedTable
    - sheet_name 存储于 ParsedTable.metadata["sheet_name"]
    - 空 Sheet 跳过
    - 旧版 XLS 格式拒绝（建议转换为 XLSX）
    - 文件大小上限保护
    """

    _XLS_MIME = "application/vnd.ms-excel"

    def parse(self, file_path: str, mime_type: str) -> ParsedDocument:
        """解析 XLSX 文件

        Args:
            file_path: 本地 XLSX 文件路径
            mime_type: MIME 类型（用于 XLS 格式拒绝）

        Returns:
            结构化解析结果
        """
        doc_id = str(uuid.uuid4())
        timestamp = datetime.now(UTC).isoformat()

        if mime_type == self._XLS_MIME:
            return ParsedDocument(
                document_id=doc_id,
                mime_type=mime_type,
                parse_status="failed",
                error_message="不支持旧版 XLS 格式，请转换为 XLSX",
                parse_timestamp=timestamp,
            )

        try:
            file_size = os.path.getsize(file_path)
        except OSError:
            logger.exception("XLSX 文件大小检查失败")
            return ParsedDocument(
                document_id=doc_id,
                mime_type=mime_type,
                parse_status="failed",
                error_message="无法访问文件，请检查文件路径或权限",
                parse_timestamp=timestamp,
            )

        if file_size == 0:
            return ParsedDocument(
                document_id=doc_id,
                mime_type=mime_type,
                parse_status="failed",
                error_message="XLSX 文档为空",
                parse_timestamp=timestamp,
            )

        if file_size > MAX_XLSX_BYTES:
            return ParsedDocument(
                document_id=doc_id,
                mime_type=mime_type,
                parse_status="failed",
                error_message=f"XLSX 文件大小 {file_size // (1024 * 1024)}MB 超过 {MAX_XLSX_BYTES // (1024 * 1024)}MB 限制",
                parse_timestamp=timestamp,
            )

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)

            tables: list[ParsedTable] = []
            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows_data: list[list[str]] = []
                    for row in ws.iter_rows(values_only=True):
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        rows_data.append(row_data)

                    if rows_data:
                        tables.append(
                            ParsedTable(
                                rows=rows_data,
                                metadata={"sheet_name": sheet_name},
                            )
                        )
            finally:
                wb.close()

            if not tables:
                return ParsedDocument(
                    document_id=doc_id,
                    mime_type=mime_type,
                    parse_status="failed",
                    error_message="XLSX 文档为空，所有 Sheet 均无数据",
                    parse_timestamp=timestamp,
                )

            page = ParsedPage(page_number=1, tables=tables)
            return ParsedDocument(
                document_id=doc_id,
                mime_type=mime_type,
                pages=[page],
                parse_status="completed",
                parse_timestamp=timestamp,
            )

        except Exception:
            logger.exception("XLSX 解析失败")
            return ParsedDocument(
                document_id=doc_id,
                mime_type=mime_type,
                parse_status="failed",
                error_message="XLSX 文档解析失败，文件可能已损坏或格式不正确",
                parse_timestamp=timestamp,
            )
